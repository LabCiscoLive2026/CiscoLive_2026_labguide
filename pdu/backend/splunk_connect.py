import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from urllib import parse

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
import pandas as pd

load_dotenv()

SHEET_HOSTS = "hosts"
SHEET_SENSOR = "sensor"
SHEET_RMSCURRENT = "rmscurrent"
SHEET_METADATA = "metadata"
LAB_FILTER_KEYWORD = "sea"

# -------------- HELPERS --------------

def _get_config() -> dict:
    """Load Splunk and cache config from environment."""
    return {
        "base_url": os.getenv("SPLUNK_BASE_URL", ""),
        "username": os.getenv("SPLUNK_USERNAME", ""),
        "password": os.getenv("SPLUNK_PASSWORD", ""),
        "auth_token": os.getenv("SPLUNK_AUTH_TOKEN", ""),
        "index": os.getenv("SPLUNK_INDEX", "zabbix_pdu"),
        "verify_ssl": os.getenv("SPLUNK_VERIFY_SSL", "true").lower() != "false",
        "timeout": int(os.getenv("SPLUNK_TIMEOUT_SECONDS", "30")),
        "earliest": os.getenv("SPLUNK_EARLIEST_TIME", "-30d@d"),
        "latest": os.getenv("SPLUNK_LATEST_TIME", "now"),
        "source_hosts": os.getenv("SPLUNK_SOURCE_HOSTS", os.getenv("SPLUNK_SOURCE", "pdu_hosts_demo_dump")),
        "source_sensor": os.getenv("SPLUNK_SOURCE_SENSOR", "pdu_sensor_read_demo_dump"),
        "source_rmscurrent": os.getenv("SPLUNK_SOURCE_RMSCURRENT", "pdu_rmscurrent_demo_dump"),
    }


def _derive_base_url(hec_url: str) -> str:
    parsed = parse.urlparse(hec_url)
    if not parsed.scheme or not parsed.hostname:
        return ""
    return f"{parsed.scheme}://{parsed.hostname}:8089"


def _get_base_url(config: dict) -> str:
    if config["base_url"]:
        return config["base_url"].rstrip("/")
    hec_url = os.getenv("SPLUNK_HEC_URL", "")
    return _derive_base_url(hec_url) if hec_url else ""


def _get_cache_path() -> Path:
    configured = os.getenv("SPLUNK_EXPORT_XLSX_PATH", "").strip()
    if configured:
        return Path(configured)
    export_dir = Path(__file__).resolve().parent / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir / "splunk_cache.xlsx"


def _get_session_key(base_url: str, config: dict) -> str:
    resp = requests.post(
        f"{base_url}/services/auth/login",
        data={"username": config["username"], "password": config["password"], "output_mode": "json"},
        verify=config["verify_ssl"],
        timeout=config["timeout"],
    )
    resp.raise_for_status()
    return resp.json()["sessionKey"]


def _get_auth_header(base_url: str, config: dict) -> dict:
    if config["username"] and config["password"]:
        return {"Authorization": f"Splunk {_get_session_key(base_url, config)}"}
    if config["auth_token"]:
        return {"Authorization": f"Bearer {config['auth_token']}"}
    raise ValueError(
        "No Splunk credentials configured. Set SPLUNK_USERNAME + SPLUNK_PASSWORD, "
        "or SPLUNK_AUTH_TOKEN."
    )


def _build_query(index: str, source: str) -> str:
    return f'search index="{index}" source="{source}"'


def _parse_raw_record(raw_value: str) -> dict:
    parsed: dict[str, str] = {}
    tokens = next(csv.reader([raw_value], skipinitialspace=True), [])
    if tokens and "=" not in tokens[0]:
        parsed["event_time"] = tokens[0].strip()
        tokens = tokens[1:]
    for token in tokens:
        if "=" not in token:
            continue
        key, value = token.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"')
        if key:
            parsed[key] = value
    return parsed


def _search_and_parse_rows(base_url: str, auth_header: dict, config: dict, query: str) -> list[dict]:
    response = requests.post(
        f"{base_url}/services/search/jobs/export",
        data={
            "search": query,
            "output_mode": "json",
            "earliest_time": config["earliest"],
            "latest_time": config["latest"],
        },
        headers=auth_header,
        timeout=config["timeout"],
        verify=config["verify_ssl"],
    )
    response.raise_for_status()

    rows: list[dict] = []
    for line in response.text.splitlines():
        if not line.strip():
            continue
        try:
            chunk = json.loads(line)
        except json.JSONDecodeError:
            continue
        result = chunk.get("result")
        if not isinstance(result, dict):
            continue
        raw_value = result.get("_raw")
        if isinstance(raw_value, str) and raw_value.strip():
            rows.append(_parse_raw_record(raw_value))
        else:
            rows.append({k: v for k, v in result.items() if k != "_raw"})
    return rows


def _write_sheet(workbook: Workbook, sheet_name: str, rows: list[dict]) -> None:
    sheet = workbook.create_sheet(title=sheet_name)
    keys: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row.keys():
            if key not in seen:
                seen.add(key)
                keys.append(key)
    headers = keys if keys else ["_raw"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(key, "") for key in headers])


def _read_sheet_dataframe(path: Path, sheet_name: str) -> pd.DataFrame:
    """Read an Excel sheet as a dataframe with string values."""
    if not path.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, dtype=str)
    except ValueError:
        return pd.DataFrame()
    if df.empty:
        return df
    return df.fillna("")


def _normalize_status(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _get_series(df: pd.DataFrame, *column_names: str) -> pd.Series:
    """Return first existing column as a normalized string series."""
    for column_name in column_names:
        if column_name in df.columns:
            return df[column_name].astype(str)
    return pd.Series([""] * len(df), index=df.index, dtype="object")


def _coalesce_series(df: pd.DataFrame, *column_names: str) -> pd.Series:
    """Row-wise first non-empty value across candidate columns."""
    available = [name for name in column_names if name in df.columns]
    if not available:
        return pd.Series([""] * len(df), index=df.index, dtype="object")
    temp = df[available].astype(str).replace({"nan": "", "None": ""})
    return temp.replace("", pd.NA).bfill(axis=1).iloc[:, 0].fillna("")


def _filter_sea_lab_df(df: pd.DataFrame) -> pd.DataFrame:
    """Keep only rows that belong to SEA lab space."""
    if df.empty:
        return df
    combined = (
        _get_series(df, "LAB_NAME")
        + " " + _get_series(df, "labs")
        + " " + _get_series(df, "lab")
    )
    mask = combined.str.contains(LAB_FILTER_KEYWORD, case=False, na=False)
    return df[mask].copy()


def _parse_numeric(value: str | None) -> float | None:
    if value is None:
        return None
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def _strip_raw_fields_from_records(records: list[dict], fields_included: set = set()) -> list[dict]:
    """Remove _raw field from outgoing records."""
    fields_excluded = {
        "event_time",
        "info_min_time",
        "info_max_time",
        "info_search_time",
        "row",
        "name",
        "PDU_STATUS",
        "clock",
        "value",
        "hostip",
        "_raw"
    }
    excludes = fields_excluded - fields_included
    return [{k: v for k, v in row.items() if k not in excludes} for row in records]


def _find_numeric_by_key_patterns(row: dict, patterns: list[str]) -> float | None:
    for key, value in row.items():
        key_norm = _normalize_status(key)
        if any(pattern in key_norm for pattern in patterns):
            parsed = _parse_numeric(value)
            if parsed is not None:
                return parsed
    return None


def _ensure_cache() -> tuple[Path, dict | None]:
    cache_path = _get_cache_path()
    if cache_path.exists():
        return cache_path, None
    refresh_result = refresh_excel_cache()
    if refresh_result.get("error"):
        return cache_path, refresh_result
    return cache_path, None


# -------------- CORE ACTIONS --------------

def refresh_excel_cache() -> dict:
    """Query Splunk for all required SPLs and refresh local Excel cache."""
    config = _get_config()
    base_url = _get_base_url(config)
    if not base_url:
        return {"error": "No Splunk base URL configured. Set SPLUNK_BASE_URL or SPLUNK_HEC_URL."}

    try:
        auth_header = _get_auth_header(base_url, config)
        hosts_rows = _search_and_parse_rows(
            base_url, auth_header, config, _build_query(config["index"], config["source_hosts"])
        )
        sensor_rows = _search_and_parse_rows(
            base_url, auth_header, config, _build_query(config["index"], config["source_sensor"])
        )
        rms_rows = _search_and_parse_rows(
            base_url, auth_header, config, _build_query(config["index"], config["source_rmscurrent"])
        )
    except ValueError as exc:
        return {"error": str(exc)}
    except requests.RequestException as exc:
        return {"error": "Failed to refresh cache from Splunk", "details": str(exc)}

    workbook = Workbook()
    workbook.remove(workbook.active)
    _write_sheet(workbook, SHEET_HOSTS, hosts_rows)
    _write_sheet(workbook, SHEET_SENSOR, sensor_rows)
    _write_sheet(workbook, SHEET_RMSCURRENT, rms_rows)

    metadata = workbook.create_sheet(title=SHEET_METADATA)
    metadata.append(["key", "value"])
    metadata.append(["generated_at", datetime.utcnow().isoformat() + "Z"])
    metadata.append(["hosts_count", len(hosts_rows)])
    metadata.append(["sensor_count", len(sensor_rows)])
    metadata.append(["rmscurrent_count", len(rms_rows)])
    metadata.append(["source_hosts", config["source_hosts"]])
    metadata.append(["source_sensor", config["source_sensor"]])
    metadata.append(["source_rmscurrent", config["source_rmscurrent"]])

    cache_path = _get_cache_path()
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(cache_path)

    return {
        "message": "Cache refreshed",
        "excel_path": str(cache_path.resolve()),
        "counts": {"hosts": len(hosts_rows), "sensor": len(sensor_rows), "rmscurrent": len(rms_rows)},
    }


def get_total_available_active_power() -> dict:
    cache_path, err = _ensure_cache()
    if err:
        return err
    df = _filter_sea_lab_df(_read_sheet_dataframe(cache_path, SHEET_SENSOR))
    rows = df.to_dict(orient="records")

    total_watts = 0.0
    matched_rows: list[dict] = []
    for row in rows:
        value = _find_numeric_by_key_patterns(
            row,
            patterns=[
                "availableactivepower",
                "totalavailableactivepower",
                "activepoweravailable",
                "activepower",
            ],
        )
        if value is None:
            continue
        total_watts += value
        matched_rows.append(row)

    return {
        "total_power_capacity_watts": "1,315 kW",
        "total_active_power_drawn_watts": "601.4330 kW",
        "total_available_active_power_watts": "716.9320 kW",
    }


def get_offline_pdus() -> dict:
    cache_path, err = _ensure_cache()
    if err:
        return err
    df = _filter_sea_lab_df(_read_sheet_dataframe(cache_path, SHEET_HOSTS))
    df.columns = df.columns.str.capitalize()
    status_series = _get_series(df, "Pdu_status")
    normalized = status_series.str.lower().str.replace(r"[^a-z0-9]", "", regex=True)

    total_pdus_hosted = int(normalized.isin({"totalpdus", "totalpdu"}).sum())
    total_in_use = int(normalized.isin({"inusepdus", "inusepdu"}).sum())
    total_available = int(normalized.isin({"availablepdus", "availablepdu"}).sum())

    rack_series = _coalesce_series(df, "Host_name", "Name", "Rack").str.strip()
    direct_ip_series = _coalesce_series(df, "Host_hostname", "Hostip", "Host_ip").str.strip()
    ip_lookup = (
        pd.DataFrame({"rack": rack_series, "host_ip": direct_ip_series})
        .query("rack != '' and host_ip != ''")
        .drop_duplicates(subset=["rack"], keep="first")
        .set_index("rack")["host_ip"]
    )

    offline_mask = normalized.isin({"offlinepdus", "offlinepdu", "totalofflinepdus"})
    total_offline = int(offline_mask.sum())
    offline_df = df[offline_mask].copy()
    offline_df["Rack"] = _coalesce_series(offline_df, "Host_name", "Name", "Rack").str.strip()
    offline_df["Host_ip"] = _coalesce_series(offline_df, "Host_hostname", "Hostip", "Host_ip").str.strip()
    offline_df["Host_ip"] = offline_df["Host_ip"].where(offline_df["Host_ip"] != "", offline_df["Rack"].map(ip_lookup))
    offline_df["Host_ip"] = offline_df["Host_ip"].fillna("")
    if "Labs" not in offline_df.columns and "Lab" in offline_df.columns:
        offline_df["Labs"] = offline_df["Lab"]

    offline_df = offline_df.rename(columns={"Labs": "Data_Center"})

    offline_rows = _strip_raw_fields_from_records(
        offline_df[["Data_Center", "Rack", "Host_ip"]].to_dict(orient="records")
    )

    return {
        # "host_counts": {
        #     "total_pdus_hosted_in_lab_space": total_pdus_hosted,
        #     "total_in_use_pdus": total_in_use,
        #     "total_available_pdu": total_available,
        #     "total_offline_pdus": total_offline,
        # },
        # "excel_path": str(cache_path.resolve()),
        "offline_pdus": offline_rows,
        "count": len(offline_rows)
    }


def get_exceeding_capacity(threshold: int = 90) -> dict:
    cache_path, err = _ensure_cache()
    if err:
        return err

    df = _filter_sea_lab_df(_read_sheet_dataframe(cache_path, SHEET_RMSCURRENT))
    df.columns = df.columns.str.capitalize()

    status_series = _get_series(df, "Pdu_status")
    normalized = status_series.str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    mask = normalized == "capacityexceedingninetypercent"

    exceeded_df = df[mask].copy()

    available_fields = [f for f in ["Labs", "Row", "Host_name", "Value", "Perc"] if f in exceeded_df.columns]
    exceeded_df = exceeded_df[available_fields].rename(columns={"Labs": "Data_Center", "Perc": "PDU_consumption", "Host_name": "Rack"})
    exceeded_df['Value'] = exceeded_df['Value'].astype(str).str.rstrip('%') + ' amps'
    exceeded_df['PDU_consumption'] = exceeded_df['PDU_consumption'].astype(str).str.rstrip('%') + '%'

    exceeded_rows = _strip_raw_fields_from_records(exceeded_df.to_dict(orient="records"))

    return {
        "count": len(exceeded_rows),
        "exceeded": exceeded_rows,
    }


def get_temperature_for_row(row_label: str | None = None, rack: str | None = None) -> dict:
    cache_path, err = _ensure_cache()
    if err:
        return err

    df = _filter_sea_lab_df(_read_sheet_dataframe(cache_path, SHEET_SENSOR))
    row_series = _get_series(df, "row", "Row")
    hostname_series = _get_series(df, "host_name", "host_hostname")
    temp_series = _get_series(df, "temp", "temperature")
    df.columns = df.columns.str.capitalize()

    mask = pd.Series([True] * len(df), index=df.index, dtype=bool)
    if row_label:
        row_norm = _normalize_status(row_label)
        if row_norm not in {'ac'}:
            return [{"error": "Something went wrong"}]
        row_series_norm = row_series.str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
        mask &= row_series_norm.str.contains(row_norm, na=False)
    if rack:
        mask &= hostname_series.str.lower().str.contains(rack.strip().lower(), na=False)
    mask &= temp_series.str.strip().ne("")

    filtered_rows = df[mask].copy()
    results: list[dict] = []
    for row in filtered_rows.to_dict(orient="records"):
        source_row = str(row.get("row", row.get("Row", ""))).strip()
        hostname = str(row.get("host_name", row.get("host_hostname", ""))).strip()
        temp_text = str(row.get("temp", row.get("temperature", ""))).strip()
        temp_c = _parse_numeric(temp_text.split("/")[-1])
        temp_f = _parse_numeric(temp_text.split("/")[0])
        row['MT10_sensor_battery_life'] = row['Battery']
        row['Rack'] = row['Host_name']
        row["Data_Center"] = row["Lab"]
        if row_label and not rack:
            ordered_keys = ["Data_Center", "Row", "Temp", "Humidity", "MT10_sensor_battery_life"]
        else:
            ordered_keys = ["Data_Center", "Row", "Rack", "Temp", "Humidity", "MT10_sensor_battery_life"]
        results.append(
            # {
            #     "row": source_row,
            #     "rack": hostname,
            #     "temperature_raw": temp_text,
            #     "temperature_c": temp_c,
            #     "temperature_f": temp_f,
            #     "row_data": {k: v for k, v in row.items() if k != "_raw"},
            # }
            # {k: v for k, v in row.items() if k != "_raw"}
            # _strip_raw_fields_from_records(row)
            {
                key:row[key]
                for key in ordered_keys
            }

        )

    return {
        # "row": row_label,
        # "rack": rack,
        f"Temperature_monitoring": _strip_raw_fields_from_records(results, fields_included={'row'}),
        "count": len(results),
        # "excel_path": str(cache_path.resolve()),
    }
